import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import scipy.signal as signal
from scipy.stats import linregress

#return first numsamples samples of data
def sample_data(magnet_data, optical_data, numsamples, scale):

    wb = load_workbook(optical_data)
    ws = wb.active

    # A1
    xcol = ws['C']
    zcol = ws['D']

    x = []
    z = []
    for timepoint in range(2, len(xcol)):
        x.append(xcol[timepoint].value)
        z.append(zcol[timepoint].value)

    peaks_optical, _ = signal.find_peaks(np.diff(x), height=[2, 15], distance=10)
    opt_samples = peaks_optical + 7
    opt_samples_values = np.array(x)[peaks_optical + 10]/scale

    # plt.plot(np.diff(x))
    # plt.plot(peaks_optical, np.diff(np.array(x))[peaks_optical], 'x')
    # plt.show()
    #
    # plt.plot(np.array(x)/scale)
    # plt.plot(opt_samples, opt_samples_values, 'x')
    # plt.show()


    high_cut = 2  # Hz
    b, a = signal.butter(4, high_cut, 'low', fs=100)
    outputs = np.load(magnet_data)[:, :, :]
    outputs = signal.filtfilt(b, a, outputs, axis=1)
    peaks_magnetic, _ = signal.find_peaks(-np.diff(outputs[0, :, 16]), height=[.001, .02], distance=100)

    t = np.arange(outputs.shape[1])/100
    print(outputs.shape[1])
    mag_samples = peaks_magnetic[0:numsamples] + 100
    mag_samples_values = outputs[0, mag_samples, 16]
    #
    # plt.plot(t[0:len(t)-1], -np.diff(outputs[0, :, 16]))
    # plt.plot(t[peaks_magnetic], -np.diff(outputs[0, :, 16])[peaks_magnetic], "x")
    # plt.xlabel("sample number")
    # plt.ylabel("x pos (mm)")
    # plt.show()
    #
    # plt.plot(t, outputs[0, :, 16])
    # plt.plot(t[mag_samples], mag_samples_values, "x")
    # plt.xlabel("sample number")
    # plt.ylabel("x pos (mm)")
    # plt.show()

    return {'opt_samples_locations' : opt_samples[0:numsamples],
            'opt_samples_values' : opt_samples_values[0:numsamples],
            'mag_samples_locations' : mag_samples[0:numsamples],
            'mag_samples_values' : mag_samples_values[0:numsamples],
            'raw_data' : np.load("f" + magnet_data)}


samples = []
for i in [1, 3, 4, 6]:
    fileid = "Dash{0}".format(i)
    samples.append(sample_data("{0}.npy".format(fileid), r"C:\Users\Nanosurface\Pictures\Camera Roll\06182022\results_2022-06-20_17-37-54\xlsx\2022-06-20_{0}_X000-reslts_user.xlsx".format(fileid), 12, 116))



for dataset in samples:
    result = linregress(dataset["opt_samples_values"], dataset["mag_samples_values"])
    print(result.rvalue**2)
    print(result.slope)
    # plt.plot(opt_samples_values, result.slope * opt_samples_values + result.intercept)
    plt.plot(dataset["opt_samples_values"] - dataset["opt_samples_values"][0], (dataset["mag_samples_values"] - dataset["mag_samples_values"][0]), 'x')
    plt.plot(dataset["opt_samples_values"] - dataset["opt_samples_values"][0], -(dataset["opt_samples_values"] - dataset["opt_samples_values"][0]))

plt.xlabel("x position optical (mm)")
plt.ylabel("magnetically predicted x position (mm)")
plt.show()

for dataset in samples:
    plt.plot((dataset["opt_samples_values"] - dataset["opt_samples_values"][0])[0:11],
             np.diff((dataset["mag_samples_values"] -
                      dataset["mag_samples_values"][0])) / np.diff(dataset["opt_samples_values"]
                                                                   - dataset["opt_samples_values"][0]) + 1)
plt.ylabel('change in predicted position w.r.t. optically measured position (mm/mm)')
plt.xlabel('optically measured displacement(mm)')
plt.show()
axes = ["X", "Y", "Z"]


for i in range(0, 3):
    for j in range(0, 3):
        for dataset in samples:
            plt.plot(dataset["opt_samples_values"],
                     dataset["raw_data"][16, j, i, dataset["mag_samples_locations"]])
            plt.xlabel('distance from origin (mm)')
            plt.ylabel('measured flux density (mT)')

        plt.title("s{0}, {1} axis".format(j, axes[i]))
        plt.show()


origin = 0
for i in range(0, 3):
    for j in range(0, 3):
        for dataset in samples:
            plt.plot(dataset["opt_samples_values"] - dataset["opt_samples_values"][origin],
                     dataset["raw_data"][16, j, i, dataset["mag_samples_locations"]]
                     - dataset["raw_data"][16, j, i, dataset["mag_samples_locations"][origin]])
            plt.xlabel('distance from origin (mm)')
            plt.ylabel('measured flux density (mT)')

        plt.title("s{0}, {1} axis".format(j, axes[i]))
        plt.show()
